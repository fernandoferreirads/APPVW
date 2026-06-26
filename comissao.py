"""
comissao.py — Módulo de Cálculo de Comissão de Vendedores

Arquitetura em camadas:
  Configuração  → COMMISSION_TABLE (única fonte de verdade)
  Leitura       → load_bigbase()
  Filtro        → filter_records()
  Cálculo       → calc_commission()
  Interface     → render_comissao() → _render_kpis / _render_table / _render_charts

Lê diretamente da aba "BIGBASE" da planilha do Dashboard (dash_url).
Autenticação reutiliza o token Microsoft já armazenado em session_state.
"""

from __future__ import annotations

import base64
import time
from datetime import date, datetime
from urllib.parse import quote as _url_quote

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import requests
import streamlit as st


# ─── Tabela de Comissões (única fonte de verdade) ─────────────────────────────
# Edite APENAS aqui para alterar valores.
# Chaves são normalizadas para uppercase no lookup — sem duplicatas necessárias.

COMMISSION_TABLE: dict[str, float] = {
    "AP":               15.00,
    "GAP":              25.00,
    "FRANQUIA":         25.00,
    "SPF BASICO":       75.00,
    "SPF BÁSICO":       75.00,
    "SPF NORMAL":      100.00,
    "SPF PLUS":        150.00,
    "SEGURO VW":       275.00,
    "SEGURO CORRETORA":100.00,
    "GE 1":            250.00,
    "GE 2":            300.00,
    "GE 3":            275.00,
    "GE 4":            325.00,
    "REV PLAN":        100.00,
    "PROTEGE BAS 24":   25.00,
    "PROT BAS 24":      25.00,
    "PROTEGE BAS 36":   35.00,
    "PROT BAS 36":      35.00,
    "PROTEGE PLUS 24":  50.00,
    "PROT PLUS 24":     50.00,
    "PROTEGE PLUS 36":  60.00,
    "PROT PLUS 36":     60.00,
}

# Lookup normalizado (uppercase) — construído uma única vez no import
_COMM_NORM: dict[str, float] = {k.upper().strip(): v for k, v in COMMISSION_TABLE.items()}


def _commission(produto: str) -> float:
    """Retorna o valor de comissão de um produto (case-insensitive, 0.0 se não mapeado)."""
    if not produto or str(produto).strip() in ("", "0", "0.0"):
        return 0.0
    return _COMM_NORM.get(str(produto).upper().strip(), 0.0)


# ─── Colunas de produto na BIGBASE ────────────────────────────────────────────
# col_interna → rótulo de exibição

PRODUCT_COLS: dict[str, str] = {
    "spf":      "SPF / Seguro",
    "app":      "APP",
    "gap":      "GAP",
    "franquia": "Franquia",
    "rev_plan": "Rev Plan",
    "ge":       "GE",
    "protege":  "Protege",
}

# ─── Especificação de colunas do BIGBASE ──────────────────────────────────────
# Cada entrada: (nome_interno, [headers aceitos em uppercase], posição_fallback)
# Posições confirmadas pelo usuário:
#   G(6)=data  M(12)=spf  N(13)=app  O(14)=gap  P(15)=franquia
#   Q(16)=rev_plan  R(17)=ge  S(18)=protege  Y(24)=vendedor  Z(25)=retorno
# Deduzido pelo offset +4 em relação à planilha de contratos:
#   V(21)=tipo_veiculo (N/S)  W(22)=sempre_novo

_BIGBASE_SPEC: list[tuple[str, list[str], int | None]] = [
    # Posições baseadas na aba BASE_PAGAMENTOS (colunas A–W)
    ("proposta",         ["PROPOSTA", "N PROPOSTA", "NUM PROPOSTA"],           0),
    ("equipe",           ["EQUIPE", "LOJA", "LOJA/EQUIPE"],                    1),
    ("data_pagto",       ["DATA_PAGAMENTO","DATA PAGAMENTO","D. PAGTO",
                          "DATA PAGTO","DT PAGTO","D.PAGTO",
                          "DATA DE PAGAMENTO"],                                2),
    ("cpf_cnpj",         ["CPF_CNPJ","CPF/CNPJ","CPF","CNPJ","DOCUMENTO"],    3),
    ("cliente",          ["CLIENTE","NOME CLIENTE","RAZAO SOCIAL"],            4),
    ("valor_veiculo",    ["VALOR_VEICULO","VALOR VEICULO","VR. VEICULO",
                          "VR VEICULO","VL VEICULO"],                          5),
    ("entrada",          ["ENTRADA","VR. ENTRADA","VL ENTRADA"],               6),
    ("valor_financiado", ["VALOR_FINANCIADO","VALOR FINANCIADO",
                          "VL FINANCIADO"],                                    7),
    ("spf",              ["SPF","SPF/SEGURO","SEGURO PROT FINANCEIRA"],        8),
    ("app",              ["AP","APP","ACID PESSOAIS","ACIDENTE PESSOAL"],      9),
    ("gap",              ["GAP"],                                              10),
    ("franquia",         ["FRANQUIA","FRANQ","SEGURO FRANQUIA"],               11),
    ("rev_plan",         ["REVISAO_PLANEJADA","REVISAO PLANEJADA","REV PLAN",
                          "REV_PLAN","REVISAO","REVISÃO"],                     12),
    ("ge",               ["GE","GARANTIA","GARANTIA ESTENDIDA"],               13),
    ("protege",          ["PROTEGE","VW PROTEGE"],                             14),
    ("prazo",            ["PRAZO"],                                            15),
    ("taxa",             ["TAXA"],                                             16),
    ("tipo_veiculo",     ["N_S","N/S","N / S","TIPO","TIPO VEICULO",
                          "TIPO_VEICULO","TIPO VEI","NOVO/SEMI","NOVO / SEMI"],17),
    ("sempre_novo",      ["SEMPRE_NOVO","SEMPRE NV","SEMPRE NOVO",
                          "SEMPRE_NV"],                                        18),
    ("peso_tabela",      ["PESO_TABELA","PESO TABELA","PESO"],                 19),
    ("vendedor",         ["VENDEDOR","CONSULTOR","NOME VENDEDOR"],             20),
    ("retorno",          ["RETORNO","RETORNO3","RETORNO 3","RETORNO F&I"],     21),
    ("pontos",           ["PONTOS_POR_CONTRATO","PONTOS POR CONTRATO",
                          "PONTOS","P/CONTRATO REAL","PONTOS/CONTRATO"],       22),
]


def _build_bigbase_df(values: list[list]) -> pd.DataFrame:
    """
    Constrói DataFrame normalizado a partir dos valores brutos do BIGBASE.

    Para cada coluna interna tenta, em ordem:
      1. Localizar pelo nome do cabeçalho (case-insensitive)
      2. Usar a posição confirmada pelo usuário como fallback

    Isso garante funcionamento independente do nome real dos cabeçalhos.
    """
    if not values or len(values) < 2:
        return pd.DataFrame()

    headers_raw = [str(c).strip() for c in values[0]]
    headers_up  = [h.upper() for h in headers_raw]

    # Resolve índice de coluna para cada spec
    col_index: dict[str, int | None] = {}
    for name, aliases, fallback in _BIGBASE_SPEC:
        idx: int | None = None
        for alias in aliases:
            if alias in headers_up:
                idx = headers_up.index(alias)
                break
        if idx is None and fallback is not None and fallback < len(headers_up):
            idx = fallback
        col_index[name] = idx

    # Constrói linhas usando os índices resolvidos
    records = []
    for row in values[1:]:
        record: dict = {}
        for name, idx in col_index.items():
            if idx is not None and idx < len(row):
                record[name] = row[idx]
            else:
                record[name] = None
        records.append(record)

    return pd.DataFrame(records)


_BIGBASE_TAB        = "BASE_PAGAMENTOS"
_ABA_OUTROS_BANCOS  = "OUTROS_BANCOS"
_ABA_CONFIGURACAO   = "CONFIGURACAO"
_SPF_FINANCEIRAS_OB = {"BRADESCO", "ITAU", "ITAÚ", "SAFRA", "SANTANDER"}
_CACHE_TTL          = 300   # segundos (5 min)

_VENDEDOR_MATRICULA: dict[str, int | None] = {
    "ALBERT ALVES TORRES":                  41273,
    "CATARINA GUEDES FERNANDES":            40469,
    "CLAUDIO HENRIQUE RODRIGUES CABRAL":    40777,
    "CLINSMAN WILKE DE VASCONCELOS":        39165,
    "FRANCISCO RICLEY DE SOUSA CARVALHO":   None,
    "GABRIEL DA SILVA ALMEIDA BARBOSA":     35372,
    "HARLEN BORGES GOMES":                  None,
    "JOSE PEREIRA NEVES":                   25375,
    "LARISSA OLIVEIRA LIMA":                35887,
    "LEANDRO MATOS CABRAL":                 24969,
    "LUCAS LEONARDO DOS SANTOS ARAUJO":     None,
    "MARCUS VINICIUS RODRIGUES LOPES":      25515,
    "NEY SANTOS CERQUEIRA":                 15149,
    "RENATO MENDES ARAUJO SANTOS":          33071,
    "RODRIGO HERCULANO TORRES SANTANA":     32452,
    "SABRINA ALMEIDA VIANA":                19865,
    "YNGRID KAREN BATISTA DE FREITAS":      41681,
    "AMAURI RODRIGUES DOS SANTOS":          9250,
    "DANILO DA ROCHA NEVES":                34836,
    "FLAVIO PEREIRA DE SOUZA":              31667,
    "JAME WILLIAMS DA SILVA COSTA":         40782,
    "RODRIGO DA SILVA PAZ":                 26506,
    "THIAGO TORRES DA SILVA GOUVES":        11111,
    "ANTONINO VITORINO DE SOUSA":           24061,
    "DOUGLAS OLIVEIRA DE MORAIS":           31755,
    "EDUARDO ALVES ROQUE":                  31754,
    "EVERTON ANICESIO VELOSO":              21412,
    "GRAZIELLE SANTOS LIMA":                40833,
    "PEDRO HENRIQUE SOARES DUTRA":          28945,
    "THOMAS RAVELLI RODRIGUES DE GODOI":    41350,
}


# ─── Camada de Leitura — Graph API ────────────────────────────────────────────

def _ms_token() -> str:
    """Retorna access token válido da session_state (renova via refresh se necessário)."""
    token = st.session_state.get("_ms_token")
    exp   = st.session_state.get("_ms_token_exp", 0)
    if token and time.time() < exp - 60:
        return token

    refresh   = st.session_state.get("_ms_refresh_token")
    client_id = st.session_state.get("_comm_client_id", "")
    if refresh and client_id:
        r = requests.post(
            "https://login.microsoftonline.com/common/oauth2/v2.0/token",
            data={
                "grant_type":    "refresh_token",
                "client_id":     client_id,
                "refresh_token": refresh,
                "scope":         "https://graph.microsoft.com/Files.ReadWrite",
            },
            timeout=15,
        )
        d = r.json()
        if "access_token" in d:
            st.session_state["_ms_token"]         = d["access_token"]
            st.session_state["_ms_token_exp"]     = time.time() + d.get("expires_in", 3600)
            st.session_state["_ms_refresh_token"] = d.get("refresh_token", refresh)
            return d["access_token"]

    raise Exception("Não autenticado. Faça login nas Configurações (🔑).")


def _resolve_file(token: str, sharing_url: str) -> tuple[str, str]:
    """Resolve sharing_url → (drive_id, item_id) com cache por URL."""
    import hashlib
    url_key   = f"_comm_file_{hashlib.md5(sharing_url.encode()).hexdigest()[:12]}"
    cached    = st.session_state.get(url_key)
    if cached:
        return cached

    encoded = base64.urlsafe_b64encode(sharing_url.encode()).decode().rstrip("=")
    r = requests.get(
        f"https://graph.microsoft.com/v1.0/shares/u!{encoded}/driveItem",
        headers={"Authorization": f"Bearer {token}"},
        timeout=20,
    )
    r.raise_for_status()
    d = r.json()
    ids = (d["parentReference"]["driveId"], d["id"])
    st.session_state[url_key] = ids
    return ids


def _find_ws_id(token: str, drive_id: str, item_id: str, tab_name: str) -> str:
    """Localiza ws_id pelo nome da aba (case-insensitive fallback)."""
    cache_key = f"_comm_ws_{item_id}_{tab_name}"
    cached = st.session_state.get(cache_key)
    if cached:
        return cached

    r = requests.get(
        f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}"
        f"/workbook/worksheets",
        headers={"Authorization": f"Bearer {token}"},
        timeout=20,
    )
    r.raise_for_status()
    sheets = r.json().get("value", [])

    # Tenta match exato, depois case-insensitive
    ws_id = None
    for ws in sheets:
        if ws["name"] == tab_name:
            ws_id = ws["id"]
            break
    if ws_id is None:
        for ws in sheets:
            if ws["name"].upper() == tab_name.upper():
                ws_id = ws["id"]
                break
    if ws_id is None:
        nomes = [ws["name"] for ws in sheets]
        raise Exception(f"Aba '{tab_name}' não encontrada. Abas disponíveis: {nomes}")

    st.session_state[cache_key] = ws_id
    return ws_id


def _read_range(token: str, drive_id: str, item_id: str, ws_id: str) -> list[list]:
    """Lê usedRange da aba (reutiliza sessão Excel se ativa)."""
    hdrs = {"Authorization": f"Bearer {token}"}
    sess = st.session_state.get(f"_xl_sess_{item_id}", "")
    if sess:
        hdrs["workbook-session-id"] = sess
    r = requests.get(
        f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}"
        f"/workbook/worksheets/{_url_quote(ws_id)}/usedRange?$select=values",
        headers=hdrs,
        timeout=40,
    )
    r.raise_for_status()
    return r.json().get("values", [])


def load_bigbase(client_id: str, sharing_url: str) -> tuple[pd.DataFrame | None, str]:
    """
    Lê e normaliza a aba BIGBASE com cache de 5 minutos.
    Retorna (df, "") em sucesso, (None, "mensagem") em erro.
    """
    # Guarda client_id para renovação de token
    st.session_state["_comm_client_id"] = client_id

    cache_key = "_comm_df_bigbase"
    ts_key    = "_comm_ts_bigbase"

    cached    = st.session_state.get(cache_key)
    cached_ts = st.session_state.get(ts_key, 0)
    if cached is not None and time.time() - cached_ts < _CACHE_TTL:
        return cached, ""

    try:
        token             = _ms_token()
        drive_id, item_id = _resolve_file(token, sharing_url)
        ws_id             = _find_ws_id(token, drive_id, item_id, _BIGBASE_TAB)
        values            = _read_range(token, drive_id, item_id, ws_id)

        if not values or len(values) < 2:
            return None, "⚠️ A aba BIGBASE está vazia ou sem dados suficientes."

        # Usa mapeamento por posição (com fallback por nome de cabeçalho)
        df = _build_bigbase_df(values)

        if df.empty:
            return None, "⚠️ Não foi possível interpretar as colunas da BIGBASE."

        df = df.dropna(how="all")

        # Tipos numéricos
        for col in ("retorno", "pontos", "valor_financiado"):
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")

        # Data — aceita tanto serial Excel (número) quanto string DD/MM/YYYY
        if "data_pagto" in df.columns:
            def _parse_date(v):
                if v is None or (isinstance(v, float) and pd.isna(v)):
                    return pd.NaT
                # Excel serial number
                if isinstance(v, (int, float)):
                    try:
                        return pd.Timestamp("1899-12-30") + pd.Timedelta(days=int(v))
                    except Exception:
                        return pd.NaT
                return pd.to_datetime(str(v), dayfirst=True, errors="coerce")

            df["data_pagto"] = df["data_pagto"].apply(_parse_date)

        # Remove linhas completamente sem dados de vendedor e data
        df = df[~(df.get("vendedor", pd.Series(dtype=str)).isna()
                  & df.get("data_pagto", pd.Series(dtype="datetime64[ns]")).isna())]

        df = df.reset_index(drop=True)

        st.session_state[cache_key] = df
        st.session_state[ts_key]    = time.time()
        return df, ""

    except requests.HTTPError as exc:
        status = exc.response.status_code if exc.response is not None else 0
        msgs   = {
            401: "❌ Token expirado. Reconecte sua conta Microsoft nas Configurações.",
            403: "❌ Sem permissão de acesso ao arquivo (403).",
            404: "❌ Arquivo não encontrado (404). Verifique o link do OneDrive.",
        }
        return None, msgs.get(status, f"❌ Erro HTTP {status}.")
    except requests.ConnectionError:
        return None, "❌ Sem conexão com a internet."
    except requests.Timeout:
        return None, "❌ Timeout ao carregar BIGBASE (40 s). Tente novamente."
    except Exception as exc:
        return None, f"❌ {exc}"


def load_outros_bancos(client_id: str, sharing_url: str) -> tuple[pd.DataFrame | None, str]:
    """
    Lê aba OUTROS_BANCOS (12 colunas A-L) com cache de 5 min.
    Retorna (df, "") em sucesso, (DataFrame vazio, "") se aba não existe,
    (None, "mensagem") em erro de autenticação/acesso.
    """
    st.session_state["_comm_client_id"] = client_id

    cache_key = "_comm_df_ob"
    ts_key    = "_comm_ts_ob"

    cached    = st.session_state.get(cache_key)
    cached_ts = st.session_state.get(ts_key, 0)
    if cached is not None and time.time() - cached_ts < _CACHE_TTL:
        return cached, ""

    _OB_COLS = ["mes", "data_pagamento", "financeira", "equipe",
                "cpf_cnpj", "cliente", "valor_financiado", "spf",
                "n_s", "tipo_retorno", "vendedor", "retorno"]

    try:
        token             = _ms_token()
        drive_id, item_id = _resolve_file(token, sharing_url)
        ws_id             = _find_ws_id(token, drive_id, item_id, _ABA_OUTROS_BANCOS)
        values            = _read_range(token, drive_id, item_id, ws_id)

        if not values or len(values) < 2:
            df = pd.DataFrame(columns=_OB_COLS)
            st.session_state[cache_key] = df
            st.session_state[ts_key]    = time.time()
            return df, ""

        records = []
        for row in values[1:]:
            record = {col: (row[i] if i < len(row) else None)
                      for i, col in enumerate(_OB_COLS)}
            records.append(record)

        df = pd.DataFrame(records).dropna(how="all")

        for col in ("retorno", "valor_financiado"):
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")

        if "data_pagamento" in df.columns:
            def _parse_ob_date(v):
                if v is None or (isinstance(v, float) and pd.isna(v)):
                    return pd.NaT
                if isinstance(v, (int, float)):
                    try:
                        return pd.Timestamp("1899-12-30") + pd.Timedelta(days=int(v))
                    except Exception:
                        return pd.NaT
                return pd.to_datetime(str(v), dayfirst=True, errors="coerce")

            df["data_pagamento"] = df["data_pagamento"].apply(_parse_ob_date)

        st.session_state[cache_key] = df
        st.session_state[ts_key]    = time.time()
        return df, ""

    except requests.HTTPError as exc:
        status = exc.response.status_code if exc.response is not None else 0
        msgs   = {
            401: "❌ Token expirado. Reconecte sua conta Microsoft.",
            403: "❌ Sem permissão de acesso (403).",
            404: "❌ Arquivo não encontrado (404).",
        }
        return None, msgs.get(status, f"❌ Erro HTTP {status}.")
    except Exception as exc:
        # Aba não existe → retorna vazio sem erro
        if "não encontrada" in str(exc).lower() or "not found" in str(exc).lower():
            df = pd.DataFrame(columns=_OB_COLS)
            st.session_state[cache_key] = df
            st.session_state[ts_key]    = time.time()
            return df, ""
        return None, f"❌ {exc}"


# ─── Camada de Filtro ─────────────────────────────────────────────────────────

def filter_records(
    df: pd.DataFrame,
    vendedor: str,
    data_ini: date,
    data_fim: date,
) -> pd.DataFrame:
    """Filtra BIGBASE por vendedor (contém, case-insensitive) e período."""
    result = df.copy()

    if "data_pagto" in result.columns:
        result = result[
            result["data_pagto"].notna()
            & (result["data_pagto"] >= pd.Timestamp(data_ini))
            & (result["data_pagto"] <= pd.Timestamp(data_fim))
        ]

    if vendedor and "vendedor" in result.columns:
        vup    = vendedor.upper().strip()
        result = result[
            result["vendedor"].fillna("").str.upper().str.contains(vup, regex=False)
        ]

    return result.reset_index(drop=True)


def calc_commission_outros_bancos(
    df_ob: pd.DataFrame,
    vendedor: str,
    data_ini: date,
    data_fim: date,
) -> dict:
    """
    Filtra OUTROS_BANCOS por período e vendedor e calcula comissões.
    Retorna dict com spf_commission, retorno_commission, total_contratos.
    """
    _zero = {"spf_commission": 0.0, "retorno_commission": 0.0, "total_contratos": 0}
    if df_ob is None or df_ob.empty:
        return _zero

    df = df_ob.copy()

    if "data_pagamento" in df.columns:
        df = df[
            df["data_pagamento"].notna()
            & (df["data_pagamento"] >= pd.Timestamp(data_ini))
            & (df["data_pagamento"] <= pd.Timestamp(data_fim))
        ]

    if vendedor and "vendedor" in df.columns:
        # Bidirecional: "GRAZIELLE" casa com "GRAZIELLE SANTOS LIMA" e vice-versa.
        # OUTROS_BANCOS grava o nome curto (chave VENDEDOR_EQUIPE); BASE_PAGAMENTOS
        # usa o nome completo extraído do PDF. O contains simples não funcionaria.
        vup = vendedor.upper().strip()
        df = df[df["vendedor"].fillna("").str.upper().str.strip().apply(
            lambda s: bool(s) and (s in vup or vup in s)
        )]

    if df.empty:
        return _zero

    # SPF: R$60 por contrato onde spf="SIM" E financeira qualifica
    spf_mask = (
        df.get("spf", pd.Series(dtype=str)).fillna("").str.upper().str.strip() == "SIM"
    ) & (
        df.get("financeira", pd.Series(dtype=str))
          .fillna("").str.upper().str.strip()
          .isin(_SPF_FINANCEIRAS_OB)
    )
    spf_commission = float(int(spf_mask.sum()) * 60.0)

    # Retorno: soma da coluna L
    retorno_commission = 0.0
    if "retorno" in df.columns:
        s = pd.to_numeric(df["retorno"], errors="coerce").sum()
        retorno_commission = float(s) if not pd.isna(s) else 0.0

    return {
        "spf_commission":    spf_commission,
        "retorno_commission": retorno_commission,
        "total_contratos":   len(df),
        "df_filtrado":       df.reset_index(drop=True),
    }


# ─── Camada de Cálculo de Comissão ────────────────────────────────────────────

def calc_commission(df: pd.DataFrame) -> dict:
    """
    Calcula comissão por produto a partir de um DataFrame já filtrado.

    Retorna dict com:
      por_produto     → list[dict]  (categoria, produto, qtd, unit, total)
      total_contratos → int
      total_comissao  → float
      total_retorno   → float
      total_produtos  → int
    """
    resultados: list[dict] = []

    for col_key, col_label in PRODUCT_COLS.items():
        if col_key not in df.columns:
            continue

        serie = df[col_key].copy()

        # Máscara de valores preenchidos
        validos = serie.apply(
            lambda x: bool(x)
            and not (isinstance(x, float) and pd.isna(x))
            and str(x).strip() not in ("", "0", "0.0")
        )
        if not validos.any():
            continue

        # Normaliza ANTES do groupby: strip + upper evita duplicatas por
        # variação de espaços/capitalização entre células do Excel
        serie_norm = serie[validos].apply(lambda x: str(x).strip().upper())

        for prod_str, grupo in serie_norm.groupby(serie_norm):
            if not prod_str:
                continue
            valor_unit = _commission(prod_str)
            if valor_unit == 0.0:
                continue  # ignora textos sem comissão mapeada ("SEM PRODUTO", etc.)
            qtd = len(grupo)
            resultados.append({
                "categoria": col_label,
                "produto":   prod_str,
                "qtd":       qtd,
                "unit":      valor_unit,
                "total":     qtd * valor_unit,
            })

    # Ordena por comissão total decrescente
    resultados.sort(key=lambda x: x["total"], reverse=True)

    total_retorno = 0.0
    if "retorno" in df.columns:
        s = pd.to_numeric(df["retorno"], errors="coerce").sum()
        total_retorno = float(s) if not pd.isna(s) else 0.0

    total_comissao = sum(r["total"] for r in resultados)

    return {
        "por_produto":      resultados,
        "total_contratos":  len(df),
        "total_produtos":   sum(r["qtd"] for r in resultados),
        "total_comissao":   total_comissao,           # só produtos
        "total_retorno":    total_retorno,             # só retorno de financiamento
        "total_bruto":      total_comissao + total_retorno,  # produtos + retorno
    }


# ─── Camada de Interface ──────────────────────────────────────────────────────

_VW_BLUE = "#001E50"
_PALETTE = ["#001E50","#0040B0","#00B0F0","#1EBE5D",
            "#FF6B35","#9B59B6","#F39C12","#E74C3C"]


def _gerar_xlsx(resultado: dict) -> bytes:
    """Gera XLSX de comissão com 4 abas: Resumo, Por Produto, Contratos VW, Outros Bancos."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    _AZUL  = "001E50"
    _CINZA = "F2F4F8"

    _hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    _hdr_fill  = PatternFill("solid", fgColor=_AZUL)
    _hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _tot_font  = Font(bold=True)

    def _auto_width(ws):
        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            if not col:
                continue
            ltr = get_column_letter(col[0].column)
            mx  = max((len(str(c.value or "")) for c in col if c.value is not None), default=8)
            ws.column_dimensions[ltr].width = min(mx + 4, 52)

    def _style_header(ws, n_cols):
        for c in range(1, n_cols + 1):
            cell = ws.cell(1, c)
            cell.font      = _hdr_font
            cell.fill      = _hdr_fill
            cell.alignment = _hdr_align
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

    wb  = Workbook()
    ob_spf     = resultado.get("ob_spf_commission",     0.0)
    ob_ret_val = resultado.get("ob_retorno_commission", 0.0)
    ob_tot     = ob_spf + ob_ret_val
    has_ob     = ob_tot > 0 or resultado.get("ob_total_contratos", 0) > 0
    total_geral = resultado["total_bruto"] + ob_tot

    # ── Aba 1: Resumo ─────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Resumo"
    ws1.column_dimensions["A"].width = 38
    ws1.column_dimensions["B"].width = 22

    # Título
    ws1.merge_cells("A1:B1")
    t = ws1["A1"]
    t.value     = "COMISSÃO DE VENDEDORES — FLOW F&I"
    t.font      = Font(bold=True, color="FFFFFF", size=13)
    t.fill      = PatternFill("solid", fgColor=_AZUL)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30

    _fmt_moeda = 'R$ #,##0.00'
    _linhas_res = [
        ("Vendedor",                   resultado["vendedor"],       False),
        ("Período",
         f"{resultado['data_ini'].strftime('%d/%m/%Y')} → "
         f"{resultado['data_fim'].strftime('%d/%m/%Y')}",           False),
        (None, None, False),
        ("Contratos no período (VW)",  resultado["total_contratos"],False),
        ("Produtos produzidos",        resultado["total_produtos"],  False),
        (None, None, False),
        ("Comissão de Produtos",       resultado["total_comissao"],  False),
        ("Retorno de Financiamento",   resultado["total_retorno"],   False),
        ("Subtotal Banco VW",          resultado["total_bruto"],     True),
    ]
    if has_ob:
        _linhas_res += [
            (None, None, False),
            ("Comissão SPF — Outros Bancos",     ob_spf,     False),
            ("Comissão Retorno — Outros Bancos", ob_ret_val, False),
        ]
    _linhas_res += [
        (None, None, False),
        ("TOTAL GERAL",  total_geral, True),
    ]

    for i, (label, value, bold) in enumerate(_linhas_res, start=2):
        if label is None:
            continue
        ca = ws1.cell(i, 1, value=label)
        cb = ws1.cell(i, 2, value=value)
        ca.font = Font(bold=bold)
        cb.font = Font(bold=bold)
        if label == "TOTAL GERAL":
            for cell in (ca, cb):
                cell.fill = PatternFill("solid", fgColor=_AZUL)
                cell.font = Font(bold=True, color="FFFFFF")
        if isinstance(value, float):
            cb.number_format = _fmt_moeda

    # ── Aba 2: Por Produto ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Por Produto")
    _PP_HDR = ["Categoria", "Produto", "Qtd", "Comissão Unit. (R$)", "Total (R$)"]
    ws2.append(_PP_HDR)
    _style_header(ws2, len(_PP_HDR))

    total_qtd  = 0
    total_com  = 0.0
    for row in resultado.get("por_produto", []):
        ws2.append([row["categoria"], row["produto"], row["qtd"], row["unit"], row["total"]])
        r = ws2.max_row
        ws2.cell(r, 4).number_format = _fmt_moeda
        ws2.cell(r, 5).number_format = _fmt_moeda
        total_qtd += row["qtd"]
        total_com += row["total"]

    if resultado.get("por_produto"):
        ws2.append(["", "TOTAL", total_qtd, "", total_com])
        r = ws2.max_row
        for c in range(1, 6):
            ws2.cell(r, c).font = _tot_font
            ws2.cell(r, c).fill = PatternFill("solid", fgColor=_CINZA)
        ws2.cell(r, 5).number_format = _fmt_moeda

    _auto_width(ws2)

    # ── Aba 3: Contratos VW ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("Contratos VW")
    _VW_COLS   = ["proposta", "data_pagto", "cliente", "cpf_cnpj",
                  "spf", "app", "gap", "franquia", "ge", "protege", "retorno", "pontos"]
    _VW_LABELS = ["Proposta", "Data Pagto", "Cliente", "CPF/CNPJ",
                  "SPF", "APP", "GAP", "Franquia", "GE", "Protege", "Retorno (R$)", "Pontos"]
    ws3.append(_VW_LABELS)
    _style_header(ws3, len(_VW_LABELS))

    df_vw = resultado.get("df_filtrado", pd.DataFrame())
    _idx_ret = _VW_COLS.index("retorno") + 1
    for _, row in df_vw.iterrows():
        linha = []
        for col in _VW_COLS:
            val = row.get(col)
            if col == "data_pagto" and pd.notna(val):
                val = pd.Timestamp(val).strftime("%d/%m/%Y")
            elif isinstance(val, float) and pd.isna(val):
                val = ""
            linha.append(val)
        ws3.append(linha)
        ws3.cell(ws3.max_row, _idx_ret).number_format = _fmt_moeda

    _auto_width(ws3)

    # ── Aba 4: Outros Bancos (condicional) ────────────────────────────────────
    df_ob = resultado.get("df_ob_filtrado")
    if has_ob and df_ob is not None and not (isinstance(df_ob, pd.DataFrame) and df_ob.empty):
        ws4 = wb.create_sheet("Outros Bancos")
        _OB_COLS   = ["data_pagamento", "financeira", "cliente", "cpf_cnpj",
                      "valor_financiado", "spf", "n_s", "tipo_retorno", "vendedor", "retorno"]
        _OB_LABELS = ["Data Pagamento", "Financeira", "Cliente", "CPF/CNPJ",
                      "Vr. Financiado (R$)", "SPF", "N/S", "Tipo Retorno", "Vendedor", "Retorno (R$)"]
        ws4.append(_OB_LABELS)
        _style_header(ws4, len(_OB_LABELS))

        _idx_vf  = _OB_COLS.index("valor_financiado") + 1
        _idx_ret = _OB_COLS.index("retorno") + 1
        for _, row in df_ob.iterrows():
            linha = []
            for col in _OB_COLS:
                val = row.get(col)
                if col == "data_pagamento" and pd.notna(val):
                    val = pd.Timestamp(val).strftime("%d/%m/%Y")
                elif isinstance(val, float) and pd.isna(val):
                    val = ""
                linha.append(val)
            ws4.append(linha)
            r = ws4.max_row
            ws4.cell(r, _idx_vf).number_format  = _fmt_moeda
            ws4.cell(r, _idx_ret).number_format = _fmt_moeda

        _auto_width(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _gerar_xlsx_equipe(equipe: str, resultados: list, data_ini, data_fim) -> bytes:
    """XLSX de equipe: Resumo comparativo + uma aba por vendedor (KPIs + por_produto)."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    _AZUL      = "001E50"
    _CINZA     = "F2F4F8"
    _fmt_moeda = "R$ #,##0.00"
    _hdr_font  = Font(bold=True, color="FFFFFF", size=11)
    _hdr_fill  = PatternFill("solid", fgColor=_AZUL)
    _hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _auto_width(ws):
        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            if not col:
                continue
            ltr = get_column_letter(col[0].column)
            mx  = max((len(str(c.value or "")) for c in col if c.value is not None), default=8)
            ws.column_dimensions[ltr].width = min(mx + 4, 50)

    def _titulo(ws, texto, n_cols, height=28):
        ws.merge_cells(f"A1:{chr(64 + min(n_cols, 26))}1")
        t = ws["A1"]
        t.value     = texto
        t.font      = Font(bold=True, color="FFFFFF", size=12)
        t.fill      = PatternFill("solid", fgColor=_AZUL)
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = height

    def _hdr_row(ws, row, headers):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row, c, value=h)
            cell.font      = _hdr_font
            cell.fill      = _hdr_fill
            cell.alignment = _hdr_align
        ws.row_dimensions[row].height = 22

    wb         = Workbook()
    periodo_str = (f"{data_ini.strftime('%d/%m/%Y')} → "
                   f"{data_fim.strftime('%d/%m/%Y')}")

    # ── Aba 1: Resumo Equipe ─────────────────────────────────────────────────
    ws1   = wb.active
    ws1.title = "Resumo Equipe"
    _RES_HDR = [
        "Vendedor", "Matrícula", "Contratos VW", "Produtos",
        "Comissão Produtos (R$)", "Retorno VW (R$)", "Subtotal VW (R$)",
        "SPF Outros Bancos (R$)", "Retorno Outros Bancos (R$)", "Total Geral (R$)",
    ]
    n_cols_res = len(_RES_HDR)

    _titulo(ws1, f"RELATÓRIO DE EQUIPE — {equipe} — {periodo_str}", n_cols_res)
    _hdr_row(ws1, 2, _RES_HDR)
    ws1.freeze_panes = "A3"

    tots = {"contr": 0, "prod": 0,
            "com": 0.0, "ret_vw": 0.0, "sub_vw": 0.0,
            "spf_ob": 0.0, "ret_ob": 0.0, "tg": 0.0}

    for res in resultados:
        ob_spf    = res.get("ob_spf_commission",     0.0)
        ob_ret    = res.get("ob_retorno_commission",  0.0)
        tg        = res["total_bruto"] + ob_spf + ob_ret
        matricula = _VENDEDOR_MATRICULA.get(res["vendedor"].upper().strip())
        ws1.append([
            res["vendedor"],
            matricula,
            res["total_contratos"],
            res["total_produtos"],
            res["total_comissao"],
            res["total_retorno"],
            res["total_bruto"],
            ob_spf, ob_ret, tg,
        ])
        rn = ws1.max_row
        ws1.cell(rn, 2).alignment = Alignment(horizontal="center")
        for c in range(5, n_cols_res + 1):
            ws1.cell(rn, c).number_format = _fmt_moeda
        tots["contr"]  += res["total_contratos"]
        tots["prod"]   += res["total_produtos"]
        tots["com"]    += res["total_comissao"]
        tots["ret_vw"] += res["total_retorno"]
        tots["sub_vw"] += res["total_bruto"]
        tots["spf_ob"] += ob_spf
        tots["ret_ob"] += ob_ret
        tots["tg"]     += tg

    ws1.append([
        "TOTAL", "", tots["contr"], tots["prod"],
        tots["com"], tots["ret_vw"], tots["sub_vw"],
        tots["spf_ob"], tots["ret_ob"], tots["tg"],
    ])
    rn = ws1.max_row
    for c in range(1, n_cols_res + 1):
        ws1.cell(rn, c).font = Font(bold=True, color="FFFFFF")
        ws1.cell(rn, c).fill = PatternFill("solid", fgColor=_AZUL)
    for c in range(5, n_cols_res + 1):
        ws1.cell(rn, c).number_format = _fmt_moeda
    _auto_width(ws1)

    # ── Abas por vendedor ────────────────────────────────────────────────────
    for res in resultados:
        vend_name  = res["vendedor"]
        sname      = vend_name[:31]
        for ch in ['/', '\\', '?', '*', '[', ']', ':']:
            sname = sname.replace(ch, '')
        sname = sname.strip() or "Vendedor"

        ws = wb.create_sheet(sname)
        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 22

        _titulo(ws, vend_name, 2, height=26)

        ob_spf    = res.get("ob_spf_commission",     0.0)
        ob_ret    = res.get("ob_retorno_commission",  0.0)
        ob_tot    = ob_spf + ob_ret
        has_ob    = ob_tot > 0 or res.get("ob_total_contratos", 0) > 0
        tg        = res["total_bruto"] + ob_tot
        matricula = _VENDEDOR_MATRICULA.get(vend_name.upper().strip())

        kpis = [
            ("Matrícula",               matricula if matricula else "—", False),
            ("Período",                  periodo_str,             False),
            (None, None, False),
            ("Contratos no período (VW)", res["total_contratos"], False),
            ("Produtos produzidos",       res["total_produtos"],   False),
            (None, None, False),
            ("Comissão de Produtos",     res["total_comissao"],   False),
            ("Retorno de Financiamento", res["total_retorno"],    False),
            ("Subtotal Banco VW",        res["total_bruto"],      True),
        ]
        if has_ob:
            kpis += [
                (None, None, False),
                ("Comissão SPF — Outros Bancos",     ob_spf, False),
                ("Comissão Retorno — Outros Bancos", ob_ret, False),
            ]
        kpis += [(None, None, False), ("TOTAL GERAL", tg, True)]

        row_ptr = 2
        for label, value, bold in kpis:
            if label is None:
                row_ptr += 1
                continue
            ca = ws.cell(row_ptr, 1, value=label)
            cb = ws.cell(row_ptr, 2, value=value)
            ca.font = Font(bold=bold)
            cb.font = Font(bold=bold)
            if label == "TOTAL GERAL":
                for cell in (ca, cb):
                    cell.fill = PatternFill("solid", fgColor=_AZUL)
                    cell.font = Font(bold=True, color="FFFFFF")
            if isinstance(value, float):
                cb.number_format = _fmt_moeda
            row_ptr += 1

        # Por Produto section
        row_ptr += 1
        ws.cell(row_ptr, 1, value="POR PRODUTO").font = Font(bold=True, size=11)
        row_ptr += 1

        _PP_HDR = ["Categoria", "Produto", "Qtd", "Comissão Unit. (R$)", "Total (R$)"]
        _hdr_row(ws, row_ptr, _PP_HDR)
        for c in ("C", "D", "E"):
            ws.column_dimensions[c].width = 10 if c == "C" else 22
        row_ptr += 1

        pp_rows = res.get("por_produto", [])
        for pp in pp_rows:
            ws.cell(row_ptr, 1, value=pp["categoria"])
            ws.cell(row_ptr, 2, value=pp["produto"])
            ws.cell(row_ptr, 3, value=pp["qtd"])
            ws.cell(row_ptr, 4, value=pp["unit"]).number_format = _fmt_moeda
            ws.cell(row_ptr, 5, value=pp["total"]).number_format = _fmt_moeda
            row_ptr += 1

        if not pp_rows:
            ws.cell(row_ptr, 1, value="Sem produtos com comissão no período")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _render_kpis(summary: dict) -> None:
    # ── Linha 1: quadro financeiro principal ──────────────────────────────────
    st.markdown("""
    <style>
    .comm-card {
        background: #f8faff;
        border: 1.5px solid #dde3ef;
        border-radius: 10px;
        padding: 1.1rem 1.4rem;
        text-align: center;
    }
    .comm-card .label {
        color: #6b7280;
        font-size: 0.78rem;
        font-weight: 500;
        letter-spacing: 0.4px;
        text-transform: uppercase;
        margin-bottom: 6px;
    }
    .comm-card .value {
        color: #001e50;
        font-size: 1.45rem;
        font-weight: 700;
        letter-spacing: -0.5px;
    }
    .comm-card.highlight {
        background: linear-gradient(135deg, #001e50 0%, #0040b0 100%);
        border-color: #001e50;
    }
    .comm-card.highlight .label { color: rgba(255,255,255,0.7); }
    .comm-card.highlight .value { color: #ffffff; font-size: 1.6rem; }
    </style>
    """, unsafe_allow_html=True)

    ob_spf   = summary.get("ob_spf_commission", 0.0)
    ob_ret   = summary.get("ob_retorno_commission", 0.0)
    ob_total = ob_spf + ob_ret
    has_ob   = ob_total > 0 or summary.get("ob_total_contratos", 0) > 0

    # ── Linha 1: Banco VW ─────────────────────────────────────────────────────
    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown(f"""
        <div class="comm-card">
            <div class="label">💼 Comissão de Produtos</div>
            <div class="value">R$ {summary['total_comissao']:,.2f}</div>
        </div>""", unsafe_allow_html=True)
    with c2:
        st.markdown(f"""
        <div class="comm-card">
            <div class="label">📈 Retorno de Financiamento</div>
            <div class="value">R$ {summary['total_retorno']:,.2f}</div>
        </div>""", unsafe_allow_html=True)
    with c3:
        if has_ob:
            st.markdown(f"""
            <div class="comm-card">
                <div class="label">🏦 Subtotal Banco VW</div>
                <div class="value">R$ {summary['total_bruto']:,.2f}</div>
            </div>""", unsafe_allow_html=True)
        else:
            st.markdown(f"""
            <div class="comm-card highlight">
                <div class="label">⭐ Total Bruto (Produtos + Retorno)</div>
                <div class="value">R$ {summary['total_bruto']:,.2f}</div>
            </div>""", unsafe_allow_html=True)

    # ── Linha 2: Outros Bancos (condicional) ──────────────────────────────────
    if has_ob:
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        ob1, ob2, ob3 = st.columns(3)
        with ob1:
            st.markdown(f"""
            <div class="comm-card">
                <div class="label">🏦 Comissão SPF — Outros Bancos</div>
                <div class="value">R$ {ob_spf:,.2f}</div>
            </div>""", unsafe_allow_html=True)
        with ob2:
            st.markdown(f"""
            <div class="comm-card">
                <div class="label">💰 Comissão Retorno — Outros Bancos</div>
                <div class="value">R$ {ob_ret:,.2f}</div>
            </div>""", unsafe_allow_html=True)
        with ob3:
            total_geral = summary["total_bruto"] + ob_total
            st.markdown(f"""
            <div class="comm-card highlight">
                <div class="label">⭐ Total Geral (VW + Outros Bancos)</div>
                <div class="value">R$ {total_geral:,.2f}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

    # ── Linha final: contadores operacionais ──────────────────────────────────
    with st.container(border=True):
        cc1, cc2 = st.columns(2)
        cc1.metric("📄 Contratos no período", f"{summary['total_contratos']:,}")
        cc2.metric("📦 Produtos produzidos",  f"{summary['total_produtos']:,}")


def _render_table(summary: dict) -> None:
    rows = summary["por_produto"]
    if not rows:
        st.info("ℹ️ Nenhum produto com comissão mapeada encontrado no período.")
        return

    df = pd.DataFrame(rows)

    # Tabela formatada
    display = pd.DataFrame({
        "Categoria":      df["categoria"],
        "Produto":        df["produto"],
        "Qtd":            df["qtd"],
        "Comissão Unit.": df["unit"].apply(lambda x: f"R$ {x:,.2f}"),
        "Total":          df["total"].apply(lambda x: f"R$ {x:,.2f}"),
    })
    st.dataframe(display, use_container_width=True, hide_index=True)

    # Totalizador
    t1, t2, t3 = st.columns(3)
    t1.metric("Tipos distintos",  len(df))
    t2.metric("Total de itens",   f"{int(df['qtd'].sum()):,}")
    t3.metric("Total comissão",   f"R$ {df['total'].sum():,.2f}")


def _render_charts(summary: dict) -> None:
    rows = summary["por_produto"]
    if not rows:
        return

    df = pd.DataFrame(rows)

    col_l, col_r = st.columns(2)

    with col_l, st.container(border=True):
        st.markdown("**Comissão por Produto (R$)**")
        fig = px.bar(
            df.sort_values("total"),
            x="total", y="produto", orientation="h",
            color="categoria",
            color_discrete_sequence=_PALETTE,
            labels={"total": "R$", "produto": "", "categoria": ""},
            text=df.sort_values("total")["total"].apply(lambda x: f"R$ {x:,.0f}"),
        )
        fig.update_traces(textposition="outside")
        fig.update_layout(
            template="plotly_white", height=max(300, len(df) * 36),
            margin=dict(l=10, r=80, t=20, b=10),
            showlegend=True, legend=dict(orientation="h", y=-0.15),
        )
        st.plotly_chart(fig, use_container_width=True)

    with col_r, st.container(border=True):
        st.markdown("**Quantidade por Produto**")
        fig2 = px.bar(
            df.sort_values("qtd", ascending=False),
            x="produto", y="qtd",
            color="categoria",
            color_discrete_sequence=_PALETTE,
            labels={"qtd": "Qtd", "produto": "", "categoria": ""},
            text="qtd",
        )
        fig2.update_traces(textposition="outside")
        fig2.update_layout(
            template="plotly_white", height=max(300, len(df) * 36),
            margin=dict(l=10, r=10, t=20, b=60),
            showlegend=False, xaxis_tickangle=-35,
        )
        st.plotly_chart(fig2, use_container_width=True)

    # Donut — participação por categoria
    if len(df["categoria"].unique()) > 1:
        cat_df = df.groupby("categoria", as_index=False)["total"].sum()
        with st.container(border=True):
            st.markdown("**Participação por Categoria**")
            col_a, col_b = st.columns([1, 2])
            with col_a:
                fig3 = go.Figure(go.Pie(
                    labels=cat_df["categoria"],
                    values=cat_df["total"],
                    hole=0.5,
                    marker=dict(colors=_PALETTE[:len(cat_df)]),
                    textinfo="label+percent",
                ))
                fig3.update_layout(
                    template="plotly_white", height=260,
                    margin=dict(l=10, r=10, t=20, b=10),
                    showlegend=False,
                )
                st.plotly_chart(fig3, use_container_width=True)
            with col_b:
                cat_display = cat_df.copy()
                cat_display["Total"] = cat_display["total"].apply(lambda x: f"R$ {x:,.2f}")
                cat_display["Part. %"] = (
                    cat_display["total"] / cat_display["total"].sum() * 100
                ).apply(lambda x: f"{x:.1f}%")
                st.dataframe(
                    cat_display[["categoria", "Total", "Part. %"]].rename(
                        columns={"categoria": "Categoria"}
                    ),
                    use_container_width=True,
                    hide_index=True,
                )


def _render_contratos_ob(df_ob: pd.DataFrame) -> None:
    """Tabela dos contratos de Outros Bancos do período filtrado."""
    if df_ob is None or df_ob.empty:
        st.info("ℹ️ Nenhum contrato de Outros Bancos no período selecionado.")
        return

    cols_show = [c for c in [
        "data_pagamento", "financeira", "cliente", "cpf_cnpj",
        "valor_financiado", "spf", "n_s", "tipo_retorno", "vendedor", "retorno",
    ] if c in df_ob.columns]

    display = df_ob[cols_show].copy()
    if "data_pagamento" in display.columns:
        display["data_pagamento"] = display["data_pagamento"].dt.strftime("%d/%m/%Y")
    for col in ("valor_financiado", "retorno"):
        if col in display.columns:
            display[col] = display[col].apply(
                lambda x: f"R$ {x:,.2f}" if pd.notna(x) and x != "" else ""
            )

    display.columns = [c.replace("_", " ").upper() for c in display.columns]
    st.dataframe(display, use_container_width=True, hide_index=True)


def _render_contratos(df_filtrado: pd.DataFrame) -> None:
    """Tabela dos contratos individuais do período filtrado."""
    cols_show = [c for c in ["proposta", "data_pagto", "cliente", "cpf_cnpj",
                              "spf", "app", "gap", "franquia", "ge", "protege",
                              "retorno", "pontos"] if c in df_filtrado.columns]
    if not cols_show:
        return

    display = df_filtrado[cols_show].copy()
    if "data_pagto" in display.columns:
        display["data_pagto"] = display["data_pagto"].dt.strftime("%d/%m/%Y")
    if "retorno" in display.columns:
        display["retorno"] = display["retorno"].apply(
            lambda x: f"R$ {x:,.2f}" if not pd.isna(x) else ""
        )

    display.columns = [c.replace("_", " ").upper() for c in display.columns]
    st.dataframe(display, use_container_width=True, hide_index=True)


# ─── Entry point ──────────────────────────────────────────────────────────────

def render_comissao(client_id: str = "", sharing_url: str = "") -> None:
    """Ponto de entrada da aba Comissão — chamado pelo app.py."""
    st.markdown("""
    <div class="section-title">
        <svg width="22" height="22" viewBox="0 0 24 24" fill="none"
             stroke="#001e50" stroke-width="2.2" stroke-linecap="round" stroke-linejoin="round">
            <line x1="12" y1="1" x2="12" y2="23"/>
            <path d="M17 5H9.5a3.5 3.5 0 0 0 0 7h5a3.5 3.5 0 0 1 0 7H6"/>
        </svg>
        <span>Comissão de Vendedores</span>
    </div>
    """, unsafe_allow_html=True)

    # ── Pré-requisitos ─────────────────────────────────────────────────────────
    if st.session_state.get("_msal_auth_status") != "authenticated":
        st.info("🔑 Faça login com sua conta Microsoft nas **Configurações** para usar esta funcionalidade.")
        return
    if not sharing_url:
        st.info("⚙️ Configure o **Link do Excel — Dashboard** nas **Configurações**. "
                "O BIGBASE deve estar nesse arquivo.")
        return

    # ── Carrega BIGBASE ────────────────────────────────────────────────────────
    with st.spinner("⏳ Carregando BIGBASE…"):
        df_base, err = load_bigbase(client_id, sharing_url)

    if err:
        st.error(err)
        col_retry, _ = st.columns([1, 4])
        with col_retry:
            if st.button("🔄 Tentar novamente", key="comm_retry"):
                st.session_state.pop("_comm_df_bigbase", None)
                st.session_state.pop("_comm_ts_bigbase", None)
                st.rerun()
        return

    if df_base is None or df_base.empty:
        st.warning("Nenhum dado encontrado na aba BIGBASE.")
        return

    ts_base = st.session_state.get("_comm_ts_bigbase")
    st.caption(
        f"BIGBASE · {len(df_base):,} registros"
        + (f" · carregado às {datetime.fromtimestamp(ts_base).strftime('%H:%M:%S')}" if ts_base else "")
    )

    # ── Filtros ────────────────────────────────────────────────────────────────
    with st.container(border=True):
        col_d1, col_d2, col_vend, col_btn = st.columns([2, 2, 4, 1])

        with col_d1:
            data_ini = st.date_input(
                "Data inicial",
                value=date.today().replace(day=1),
                key="comm_data_ini",
                format="DD/MM/YYYY",
            )
        with col_d2:
            data_fim = st.date_input(
                "Data final",
                value=date.today(),
                key="comm_data_fim",
                format="DD/MM/YYYY",
            )
        with col_vend:
            vendedores_disp = (
                sorted(df_base["vendedor"].dropna().str.strip().unique())
                if "vendedor" in df_base.columns else []
            )
            vendedor_sel = st.selectbox(
                "Vendedor",
                options=[""] + vendedores_disp,
                format_func=lambda x: "Selecione um vendedor..." if x == "" else x,
                key="comm_vendedor",
            )
        with col_btn:
            st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
            consultar = st.button(
                "🔍 Consultar",
                type="primary",
                use_container_width=True,
                key="comm_consultar",
            )

    # ── Valida e executa consulta ──────────────────────────────────────────────
    if consultar:
        if not vendedor_sel:
            st.warning("⚠️ Selecione um vendedor para consultar.")
            return
        if data_ini > data_fim:
            st.warning("⚠️ A data inicial deve ser anterior ou igual à data final.")
            return

        df_filtrado = filter_records(df_base, vendedor_sel, data_ini, data_fim)

        if df_filtrado.empty:
            st.warning(
                f"⚠️ Nenhum registro encontrado para **{vendedor_sel}** "
                f"entre **{data_ini.strftime('%d/%m/%Y')}** e "
                f"**{data_fim.strftime('%d/%m/%Y')}**."
            )
            st.session_state.pop("comm_resultado", None)
            return

        # Carrega e calcula Outros Bancos (silencioso se aba não existir)
        df_ob, ob_err = load_outros_bancos(client_id, sharing_url)
        if ob_err:
            st.warning(f"⚠️ Outros Bancos: {ob_err}")
            df_ob = pd.DataFrame()
        ob_result = calc_commission_outros_bancos(
            df_ob if df_ob is not None else pd.DataFrame(),
            vendedor_sel, data_ini, data_fim,
        )

        summary = calc_commission(df_filtrado)
        summary["vendedor"]              = vendedor_sel
        summary["data_ini"]              = data_ini
        summary["data_fim"]              = data_fim
        summary["df_filtrado"]           = df_filtrado
        summary["ob_spf_commission"]     = ob_result["spf_commission"]
        summary["ob_retorno_commission"] = ob_result["retorno_commission"]
        summary["ob_total_contratos"]    = ob_result["total_contratos"]
        summary["df_ob_filtrado"]        = ob_result.get("df_filtrado", pd.DataFrame())
        st.session_state["comm_resultado"] = summary

    # ── Relatório de Equipe ────────────────────────────────────────────────────
    st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)
    with st.expander("📊 Relatório de Equipe — Exportar XLSX", expanded=False):
        with st.container(border=True):
            col_eq, col_eq1, col_eq2, col_eq_btn = st.columns([2, 2, 2, 1])
            with col_eq:
                _equipes_disp = (
                    sorted(df_base["equipe"].dropna().str.strip().str.upper().unique())
                    if "equipe" in df_base.columns else []
                )
                equipe_sel_eq = st.selectbox(
                    "Equipe",
                    options=[""] + _equipes_disp,
                    format_func=lambda x: "Selecione uma equipe..." if x == "" else x,
                    key="comm_equipe_sel",
                )
            with col_eq1:
                eq_ini = st.date_input(
                    "Data inicial",
                    value=date.today().replace(day=1),
                    key="comm_eq_ini",
                    format="DD/MM/YYYY",
                )
            with col_eq2:
                eq_fim = st.date_input(
                    "Data final",
                    value=date.today(),
                    key="comm_eq_fim",
                    format="DD/MM/YYYY",
                )
            with col_eq_btn:
                st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                gerar_eq = st.button(
                    "📊 Gerar",
                    key="comm_eq_gerar",
                    type="primary",
                    use_container_width=True,
                )

        if gerar_eq:
            if not equipe_sel_eq:
                st.warning("⚠️ Selecione uma equipe.")
            elif eq_ini > eq_fim:
                st.warning("⚠️ Data inicial deve ser anterior à data final.")
            else:
                mask_eq = (
                    df_base["equipe"].fillna("").str.strip().str.upper()
                    == equipe_sel_eq.upper()
                )
                vendedores_eq = sorted(
                    df_base[mask_eq]["vendedor"].dropna().str.strip().unique()
                ) if "equipe" in df_base.columns else []

                if not vendedores_eq:
                    st.warning(f"⚠️ Nenhum vendedor encontrado para **{equipe_sel_eq}**.")
                else:
                    df_ob_eq, ob_err_eq = load_outros_bancos(client_id, sharing_url)
                    if ob_err_eq:
                        df_ob_eq = pd.DataFrame()

                    resultados_eq = []
                    with st.spinner(
                        f"⏳ Processando {len(vendedores_eq)} vendedores…"
                    ):
                        for vend in vendedores_eq:
                            df_v    = filter_records(df_base, vend, eq_ini, eq_fim)
                            summ_v  = calc_commission(df_v)
                            ob_v    = calc_commission_outros_bancos(
                                df_ob_eq if df_ob_eq is not None else pd.DataFrame(),
                                vend, eq_ini, eq_fim,
                            )
                            summ_v["vendedor"]              = vend
                            summ_v["data_ini"]              = eq_ini
                            summ_v["data_fim"]              = eq_fim
                            summ_v["df_filtrado"]           = df_v
                            summ_v["ob_spf_commission"]     = ob_v["spf_commission"]
                            summ_v["ob_retorno_commission"] = ob_v["retorno_commission"]
                            summ_v["ob_total_contratos"]    = ob_v["total_contratos"]
                            summ_v["df_ob_filtrado"]        = ob_v.get("df_filtrado", pd.DataFrame())
                            resultados_eq.append(summ_v)

                    xlsx_eq  = _gerar_xlsx_equipe(equipe_sel_eq, resultados_eq, eq_ini, eq_fim)
                    nome_eq  = (
                        f"Comissao_Equipe_{equipe_sel_eq.replace(' ', '_')}_"
                        f"{eq_ini.strftime('%d%m%Y')}-{eq_fim.strftime('%d%m%Y')}.xlsx"
                    )
                    st.session_state["comm_eq_xlsx"]   = xlsx_eq
                    st.session_state["comm_eq_nome"]   = nome_eq
                    st.session_state["comm_eq_nvend"]  = len(vendedores_eq)

        eq_xlsx = st.session_state.get("comm_eq_xlsx")
        if eq_xlsx:
            col_dl, _ = st.columns([1, 4])
            with col_dl:
                st.download_button(
                    f"📥 Baixar ({st.session_state.get('comm_eq_nvend', 0)} vendedores)",
                    data=eq_xlsx,
                    file_name=st.session_state.get("comm_eq_nome", "relatorio_equipe.xlsx"),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="comm_eq_download",
                )

    # ── Exibe resultado ────────────────────────────────────────────────────────
    resultado = st.session_state.get("comm_resultado")
    if not resultado:
        st.info("Selecione o período e o vendedor, depois clique em **🔍 Consultar**.")
        return

    st.divider()

    # Cabeçalho do resultado
    st.markdown(
        f"### {resultado['vendedor']} &nbsp;·&nbsp; "
        f"{resultado['data_ini'].strftime('%d/%m/%Y')} → "
        f"{resultado['data_fim'].strftime('%d/%m/%Y')}"
    )

    # KPIs
    _render_kpis(resultado)

    # Abas internas: Detalhamento | Contratos | Outros Bancos
    sub_det, sub_con, sub_ob = st.tabs([
        "📋 Detalhamento por Produto",
        "📄 Contratos do Período",
        "🏦 Outros Bancos",
    ])

    with sub_det:
        st.markdown("#### Comissão por produto")
        _render_table(resultado)
        st.divider()
        st.markdown("#### Gráficos de produção")
        _render_charts(resultado)

    with sub_con:
        st.markdown(
            f"**{resultado['total_contratos']} contrato(s)** no período filtrado"
        )
        _render_contratos(resultado["df_filtrado"])

    with sub_ob:
        df_ob_disp = resultado.get("df_ob_filtrado")
        n_ob       = resultado.get("ob_total_contratos", 0)

        # Detecta resultado desatualizado: contagem > 0 mas DataFrame ausente/vazio
        dados_incompletos = (
            df_ob_disp is None
            or (isinstance(df_ob_disp, pd.DataFrame) and df_ob_disp.empty and n_ob > 0)
        )
        if dados_incompletos:
            st.warning(
                "⚠️ Dados desatualizados. Clique em **🔍 Consultar** novamente "
                "para exibir os contratos de Outros Bancos."
            )
        else:
            n_real = len(df_ob_disp) if df_ob_disp is not None else 0
            st.markdown(f"**{n_real} contrato(s) de Outros Bancos** no período filtrado")
            _render_contratos_ob(df_ob_disp if df_ob_disp is not None else pd.DataFrame())

    # Botões de ação
    col_lim, col_att, col_exp, _ = st.columns([1, 1, 1, 3])
    with col_lim:
        if st.button("🗑️ Limpar", key="comm_clear", use_container_width=True):
            st.session_state.pop("comm_resultado", None)
            st.rerun()
    with col_att:
        if st.button("🔄 Recarregar base", key="comm_reload", use_container_width=True):
            st.session_state.pop("_comm_df_bigbase", None)
            st.session_state.pop("_comm_ts_bigbase", None)
            st.session_state.pop("_comm_df_ob", None)
            st.session_state.pop("_comm_ts_ob", None)
            st.session_state.pop("comm_resultado", None)
            st.rerun()
    with col_exp:
        xlsx_bytes = _gerar_xlsx(resultado)
        nome_arq = (
            f"Comissao_{resultado['vendedor'].replace(' ', '_')}_"
            f"{resultado['data_ini'].strftime('%d%m%Y')}-"
            f"{resultado['data_fim'].strftime('%d%m%Y')}.xlsx"
        )
        st.download_button(
            "📥 Exportar XLSX",
            data=xlsx_bytes,
            file_name=nome_arq,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="comm_export",
        )


# ─── Período de Fechamento (aba CONFIGURACAO) ─────────────────────────────────

def load_periodo_fechamento(
    client_id: str, sharing_url: str
) -> tuple[date | None, date | None]:
    """Lê data_ini e data_fim da aba CONFIGURACAO. Retorna (None, None) se não configurado."""
    cache_key = "_comm_periodo_fechamento"
    ts_key    = "_comm_ts_periodo"
    cached    = st.session_state.get(cache_key)
    cached_ts = st.session_state.get(ts_key, 0)
    if cached is not None and time.time() - cached_ts < _CACHE_TTL:
        return cached

    st.session_state["_comm_client_id"] = client_id
    try:
        token             = _ms_token()
        drive_id, item_id = _resolve_file(token, sharing_url)
        ws_id             = _find_ws_id(token, drive_id, item_id, _ABA_CONFIGURACAO)
        values            = _read_range(token, drive_id, item_id, ws_id)

        data_map: dict[str, str] = {}
        for row in values:
            if len(row) >= 2 and row[0] and row[1]:
                data_map[str(row[0]).lower().strip()] = str(row[1]).strip()

        ini_str = data_map.get("periodo_ini")
        fim_str = data_map.get("periodo_fim")

        ini = date.fromisoformat(ini_str) if ini_str else None
        fim = date.fromisoformat(fim_str) if fim_str else None

        result = (ini, fim)
        st.session_state[cache_key] = result
        st.session_state[ts_key]    = time.time()
        return result
    except Exception:
        return None, None


def save_periodo_fechamento(
    client_id: str, sharing_url: str, data_ini: date, data_fim: date
) -> str:
    """Grava data_ini e data_fim na aba CONFIGURACAO da planilha. Retorna '' em sucesso."""
    st.session_state["_comm_client_id"] = client_id
    try:
        token             = _ms_token()
        drive_id, item_id = _resolve_file(token, sharing_url)

        try:
            ws_id = _find_ws_id(token, drive_id, item_id, _ABA_CONFIGURACAO)
        except Exception:
            r = requests.post(
                f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}"
                f"/workbook/worksheets",
                headers={"Authorization": f"Bearer {token}",
                         "Content-Type": "application/json"},
                json={"name": _ABA_CONFIGURACAO},
                timeout=20,
            )
            r.raise_for_status()
            ws_id = r.json()["id"]
            st.session_state.pop(f"_comm_ws_{item_id}_{_ABA_CONFIGURACAO}", None)

        valores = [
            ["chave",        "valor"],
            ["periodo_ini",  data_ini.isoformat()],
            ["periodo_fim",  data_fim.isoformat()],
        ]
        r = requests.patch(
            f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}"
            f"/workbook/worksheets/{_url_quote(ws_id)}/range(address='A1:B3')",
            headers={"Authorization": f"Bearer {token}",
                     "Content-Type": "application/json"},
            json={"values": valores},
            timeout=20,
        )
        r.raise_for_status()

        st.session_state.pop("_comm_periodo_fechamento", None)
        st.session_state.pop("_comm_ts_periodo", None)
        return ""
    except Exception as exc:
        return f"❌ {exc}"


# ─── Painel Individual do Vendedor ────────────────────────────────────────────

def render_painel_vendedor(client_id: str = "", sharing_url: str = "") -> None:
    """Aba 'Minha Produção' — login por matrícula + visão individual de comissão."""

    if not client_id or not sharing_url:
        st.info("⚙️ Configure o Client ID e o link do Excel nas Configurações.")
        return

    vend_logado: str | None = st.session_state.get("_vend_logado")

    # ── Tela de login ─────────────────────────────────────────────────────────
    if not vend_logado:
        st.markdown(
            "<p style='font-size:1.1rem;font-weight:700;color:#001e50;"
            "margin-bottom:1rem'>👤 Acesso ao Painel do Vendedor</p>",
            unsafe_allow_html=True,
        )
        with st.container(border=True):
            vendedores_lista = sorted(_VENDEDOR_MATRICULA.keys())
            nome_sel = st.selectbox(
                "Seu nome",
                options=[""] + vendedores_lista,
                format_func=lambda x: "Selecione seu nome..." if x == "" else x,
                key="vend_nome_login",
            )
            pin_input = st.number_input(
                "Matrícula",
                min_value=0, max_value=999999,
                value=0, step=1,
                key="vend_pin_login",
            )
            if st.button("🔓 Entrar", use_container_width=True, key="vend_btn_login"):
                if not nome_sel:
                    st.error("Selecione seu nome.")
                else:
                    mat = _VENDEDOR_MATRICULA.get(nome_sel)
                    if mat is None:
                        st.warning("⚠️ Matrícula não cadastrada ainda. Procure o gestor.")
                    elif int(pin_input) == int(mat):
                        st.session_state["_vend_logado"] = nome_sel
                        st.rerun()
                    else:
                        st.error("❌ Matrícula incorreta. Tente novamente.")
        return

    # ── Painel logado ─────────────────────────────────────────────────────────
    col_h, col_sair = st.columns([7, 1])
    col_h.markdown(
        f"<p style='font-size:1rem;font-weight:700;color:#001e50;margin:0'>"
        f"👤 {vend_logado}</p>",
        unsafe_allow_html=True,
    )
    if col_sair.button("🚪 Sair", key="vend_btn_sair", use_container_width=True):
        st.session_state.pop("_vend_logado", None)
        st.rerun()

    # Carrega período de fechamento
    data_ini, data_fim = load_periodo_fechamento(client_id, sharing_url)
    if data_ini is None or data_fim is None:
        st.warning(
            "⚠️ Período de fechamento ainda não configurado. "
            "Aguarde o gestor definir o período nas Configurações."
        )
        return

    st.info(
        f"📅 Período de fechamento: "
        f"**{data_ini.strftime('%d/%m/%Y')} → {data_fim.strftime('%d/%m/%Y')}**"
    )

    # Carrega dados da base
    st.session_state["_comm_client_id"] = client_id
    df_base, err = load_bigbase(client_id, sharing_url)
    if err:
        st.error(err)
        return
    if df_base is None or df_base.empty:
        st.warning("⚠️ Base de dados vazia ou inacessível.")
        return

    # Filtra e calcula comissão
    df_filtrado = filter_records(df_base, vend_logado, data_ini, data_fim)
    summary     = calc_commission(df_filtrado)

    df_ob, _ = load_outros_bancos(client_id, sharing_url)
    ob_result = calc_commission_outros_bancos(
        df_ob if df_ob is not None else pd.DataFrame(),
        vend_logado, data_ini, data_fim,
    )

    summary["vendedor"]              = vend_logado
    summary["data_ini"]              = data_ini
    summary["data_fim"]              = data_fim
    summary["ob_spf_commission"]     = ob_result["spf_commission"]
    summary["ob_retorno_commission"] = ob_result["retorno_commission"]
    summary["ob_total_contratos"]    = ob_result["total_contratos"]
    summary["df_ob_filtrado"]        = ob_result.get("df_filtrado", pd.DataFrame())
    summary["df_filtrado"]           = df_filtrado

    _render_kpis(summary)

    sub_det, sub_con, sub_ob = st.tabs([
        "📋 Detalhamento por Produto",
        "📄 Contratos do Período",
        "🏦 Outros Bancos",
    ])

    with sub_det:
        _render_table(summary)

    with sub_con:
        _render_contratos(df_filtrado)

    with sub_ob:
        df_ob_filtrado = summary.get("df_ob_filtrado", pd.DataFrame())
        n_ob = ob_result["total_contratos"]
        if n_ob == 0:
            st.info("Nenhum contrato de Outros Bancos no período selecionado.")
        else:
            _render_contratos_ob(df_ob_filtrado)
